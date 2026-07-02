import logging
import os
from io import BytesIO
from pathlib import Path

import chardet

logger = logging.getLogger(__name__)

# Docling gives layout-aware Markdown (tables, headings, page markers) which
# unlocks the structured chunker for PDFs. Falls back to pypdf on any failure.
DOCLING_ENABLED = os.environ.get("DOCLING_ENABLED", "1") == "1"
_docling_converter = None

SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".py", ".js", ".ts", ".json", ".yaml", ".yml",
    ".xml", ".html", ".css", ".csv", ".log", ".cfg", ".ini", ".conf",
    ".sh", ".bash", ".zsh", ".bat", ".ps1", ".sql", ".r", ".go",
    ".java", ".c", ".cpp", ".h", ".hpp", ".rs", ".toml", ".zig",
    ".pdf", ".docx", ".xlsx",
}


def can_parse(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_EXTENSIONS


def parse_file(file_path: str | None = None, content: bytes | None = None, filename: str = "") -> str:
    if file_path:
        filename = filename or file_path
        with open(file_path, "rb") as f:
            content = f.read()

    if content is None:
        return ""

    ext = Path(filename).suffix.lower()

    try:
        if ext == ".pdf":
            return _parse_pdf(content)
        elif ext == ".docx":
            return _parse_docx(content)
        elif ext == ".xlsx":
            return _parse_xlsx(content)
        else:
            return _parse_text(content)
    except Exception as e:
        logger.error(f"Failed to parse {filename}: {e}")
        return ""


def _parse_text(content: bytes) -> str:
    detected = chardet.detect(content)
    encoding = detected.get("encoding", "utf-8") or "utf-8"
    try:
        return content.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        return content.decode("utf-8", errors="replace")


def _get_docling_converter():
    global _docling_converter
    if _docling_converter is None:
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.document_converter import DocumentConverter, PdfFormatOption
        opts = PdfPipelineOptions(do_ocr=False)
        _docling_converter = DocumentConverter(
            format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=opts)}
        )
    return _docling_converter


def _parse_pdf_docling(content: bytes) -> str:
    from docling.datamodel.base_models import DocumentStream
    converter = _get_docling_converter()
    result = converter.convert(DocumentStream(name="upload.pdf", stream=BytesIO(content)))
    doc = result.document

    # Per-page Markdown with <!-- page N --> markers so the structured chunker
    # can attach page numbers to chunks. Fall back to a single whole-doc export
    # if this docling version doesn't support page_no.
    try:
        page_nums = sorted(doc.pages.keys())
        parts = []
        for p in page_nums:
            md = doc.export_to_markdown(page_no=p)
            if md.strip():
                parts.append(f"<!-- page {p} -->\n{md}")
        if parts:
            return "\n\n".join(parts)
    except TypeError:
        pass
    return doc.export_to_markdown()


def _parse_pdf(content: bytes) -> str:
    if DOCLING_ENABLED:
        try:
            text = _parse_pdf_docling(content)
            if text.strip():
                return text
        except Exception as e:
            logger.warning("Docling parse failed, falling back to pypdf: %s", e)
    return _parse_pdf_pypdf(content)


def _parse_pdf_pypdf(content: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(BytesIO(content))
    text_parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            text_parts.append(text)
    return "\n\n".join(text_parts)


def _parse_docx(content: bytes) -> str:
    from docx import Document
    doc = Document(BytesIO(content))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), read_only=True)
    text_parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text_parts.append(f"--- Sheet: {sheet} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            text_parts.append("\t".join(cells))
    return "\n".join(text_parts)
